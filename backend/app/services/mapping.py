"""导入映射模板 CRUD + Excel 表头预览"""

from __future__ import annotations

import uuid
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import Select, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from xlrd import open_workbook

from app.core.exceptions import BadRequest, NotFound
from app.models.ingest import FieldMapping
from app.schemas.mapping import (
    HeaderMatchResult,
    MappingCreate,
    MappingResponse,
    MappingUpdate,
    PreviewResponse,
)
from app.utils.db import active_filter, paginate


# ── 系统字段列表（用于预览界面展示） ────────────────

SYSTEM_FIELDS: dict[str, str] = {
    "data_date": "数据日期",
    "category": "分部工程类别",
    "item_name": "分项名称",
    "planned_quantity": "计划工程量",
    "actual_quantity": "实际工程量",
    "unit": "单位",
    "unit_price": "单价（分）",
    "amount": "金额（分）",
    "cost_type": "费用类型",
}

SYSTEM_CONVERTERS: dict[str, str] = {
    "wan_yuan_to_fen": "万元转分",
    "yuan_to_fen": "元转分",
    "excel_serial_to_date": "Excel 日期序列号转日期",
    "chinese_date_to_iso": "中文日期转 ISO 日期",
    "trim": "去除首尾空格",
    "lowercase": "转小写",
    "category_alias_to_canonical": "施工类别别名规范化",
    "decimal_to_base_unit": "工程量单位换算",
}


# ── 预置模板 ────────────────────────────────────────

PRESET_TEMPLATES: dict[str, dict] = {
    "weekly": {
        "mapping_name": "周报标准模板",
        "biz_type": "weekly",
        "file_format": "xlsx",
        "header_row": 1,
        "sheet_index": 0,
        "is_active": True,
        "rules": [
            {"user_header": "日期", "system_field": "data_date", "converter": "excel_serial_to_date"},
            {"user_header": "分部工程", "system_field": "category", "converter": ""},
            {"user_header": "分项名称", "system_field": "item_name", "converter": ""},
            {"user_header": "单位", "system_field": "unit", "converter": ""},
            {"user_header": "计划工程量", "system_field": "planned_quantity", "converter": ""},
            {"user_header": "实际工程量", "system_field": "actual_quantity", "converter": ""},
            {"user_header": "单价(元)", "system_field": "unit_price", "converter": "yuan_to_fen"},
            {"user_header": "金额(元)", "system_field": "amount", "converter": "yuan_to_fen"},
            {"user_header": "费用类型", "system_field": "cost_type", "converter": ""},
        ],
    },
    "monthly": {
        "mapping_name": "月报标准模板",
        "biz_type": "monthly",
        "file_format": "xlsx",
        "header_row": 1,
        "sheet_index": 0,
        "is_active": True,
        "rules": [
            {"user_header": "数据日期", "system_field": "data_date", "converter": "excel_serial_to_date"},
            {"user_header": "分部工程", "system_field": "category", "converter": ""},
            {"user_header": "项目名称", "system_field": "item_name", "converter": ""},
            {"user_header": "单位", "system_field": "unit", "converter": ""},
            {"user_header": "计划量", "system_field": "planned_quantity", "converter": ""},
            {"user_header": "完成量", "system_field": "actual_quantity", "converter": ""},
            {"user_header": "单价(万元)", "system_field": "unit_price", "converter": "wan_yuan_to_fen"},
            {"user_header": "金额(万元)", "system_field": "amount", "converter": "wan_yuan_to_fen"},
        ],
    },
    "progress_payment": {
        "mapping_name": "进度款标准模板",
        "biz_type": "progress_payment",
        "file_format": "xlsx",
        "header_row": 1,
        "sheet_index": 0,
        "is_active": True,
        "rules": [
            {"user_header": "日期", "system_field": "data_date", "converter": "excel_serial_to_date"},
            {"user_header": "分部工程", "system_field": "category", "converter": ""},
            {"user_header": "项目名称", "system_field": "item_name", "converter": ""},
            {"user_header": "单位", "system_field": "unit", "converter": ""},
            {"user_header": "申报工程量", "system_field": "planned_quantity", "converter": ""},
            {"user_header": "核定工程量", "system_field": "actual_quantity", "converter": ""},
            {"user_header": "单价(元)", "system_field": "unit_price", "converter": "yuan_to_fen"},
            {"user_header": "金额(元)", "system_field": "amount", "converter": "yuan_to_fen"},
            {"user_header": "费用类型", "system_field": "cost_type", "converter": ""},
        ],
    },
}


# ── 查询 ────────────────────────────────────────────


def _query_base() -> Select:
    return select(FieldMapping).order_by(FieldMapping.created_at.desc())


async def list_mappings(
    db: AsyncSession,
    biz_type: str | None = None,
    keyword: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[FieldMapping], int]:
    """分页查询映射模板（自动过滤软删除）。"""
    stmt = _query_base()

    if biz_type:
        stmt = stmt.where(FieldMapping.biz_type == biz_type)
    if keyword:
        stmt = stmt.where(FieldMapping.mapping_name.ilike(f"%{keyword}%"))

    return await paginate(db, stmt, model=FieldMapping, page=page, page_size=page_size)


async def get_mapping(db: AsyncSession, mapping_id: uuid.UUID) -> FieldMapping:
    """获取单条映射模板（不含已软删除）。"""
    stmt = active_filter(select(FieldMapping).where(FieldMapping.id == mapping_id), FieldMapping)
    result = await db.execute(stmt)
    obj = result.scalar_one_or_none()
    if not obj:
        raise NotFound(detail="映射模板不存在")
    return obj


# ── 写入 ────────────────────────────────────────────


async def create_mapping(db: AsyncSession, data: MappingCreate) -> FieldMapping:
    """创建映射模板。"""
    obj = FieldMapping(
        mapping_name=data.mapping_name,
        biz_type=data.biz_type,
        file_format=data.file_format,
        header_row=data.header_row,
        sheet_index=data.sheet_index,
        project_id=data.project_id,
        is_active=data.is_active,
        rules={"rules": [r.model_dump() for r in data.rules]},
    )
    db.add(obj)
    await db.flush()
    await db.refresh(obj)
    return obj


async def update_mapping(db: AsyncSession, mapping_id: uuid.UUID, data: MappingUpdate) -> FieldMapping:
    """更新映射模板（部分更新）。"""
    obj = await get_mapping(db, mapping_id)
    update_data = data.model_dump(exclude_unset=True)
    if "rules" in update_data:
        update_data["rules"] = {"rules": [r.model_dump() for r in data.rules]}  # type: ignore
    for key, value in update_data.items():
        setattr(obj, key, value)
    await db.flush()
    await db.refresh(obj)
    return obj


async def delete_mapping(db: AsyncSession, mapping_id: uuid.UUID) -> None:
    """软删除映射模板。"""
    from datetime import UTC, datetime

    obj = await get_mapping(db, mapping_id)
    obj.deleted_at = datetime.now(UTC)
    await db.flush()


async def seed_presets(db: AsyncSession, override: bool = False) -> None:
    """写入预置模板（仅当对应 biz_type 不存在时）。"""
    for biz_type, preset in PRESET_TEMPLATES.items():
        if not override:
            result = await db.execute(
                select(FieldMapping.id).where(
                    FieldMapping.biz_type == biz_type,
                    FieldMapping.mapping_name == preset["mapping_name"],
                    FieldMapping.deleted_at.is_(None),
                ).limit(1)
            )
            if result.first():
                continue
        obj = FieldMapping(
            mapping_name=preset["mapping_name"],
            biz_type=preset["biz_type"],
            file_format=preset["file_format"],
            header_row=preset["header_row"],
            sheet_index=preset["sheet_index"],
            is_active=preset["is_active"],
            rules={"rules": preset["rules"]},
        )
        db.add(obj)
    await db.flush()
    logger = __import__("loguru").logger
    logger.info("预置映射模板已写入")


# ── Excel 表头预览 ─────────────────────────────────


def _read_excel_headers(file_data: bytes, ext: str, sheet_index: int = 0, header_row: int = 1) -> list[str]:
    """读取 Excel 文件表头行，返回表头名称列表。"""
    if ext == "csv":
        import csv
        from io import StringIO

        try:
            text = file_data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_data.decode("gbk")
        reader = csv.reader(StringIO(text))
        for _ in range(header_row - 1):
            next(reader, None)
        row = next(reader, None)
        if not row:
            raise BadRequest(detail="CSV 文件为空或表头行不存在")
        return [h.strip() for h in row if h.strip()]

    if ext == "xlsx":
        wb = load_workbook(BytesIO(file_data), read_only=True, data_only=True)
        if sheet_index >= len(wb.sheetnames):
            raise BadRequest(detail=f"Sheet 索引 {sheet_index} 超出范围（共 {len(wb.sheetnames)} 个）")
        ws = wb[wb.sheetnames[sheet_index]]
        for _ in range(header_row - 1):
            next(ws.iter_rows(), None)
        row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), [])
        wb.close()
        return [str(c).strip() for c in row if c is not None and str(c).strip()]

    if ext == "xls":
        wb = open_workbook(file_contents=file_data)
        if sheet_index >= wb.nsheets:
            raise BadRequest(detail=f"Sheet 索引 {sheet_index} 超出范围（共 {wb.nsheets} 个）")
        ws = wb.sheet_by_index(sheet_index)
        if header_row > ws.nrows:
            raise BadRequest(detail=f"表头行 {header_row} 超出文件总行数 {ws.nrows}")
        return [str(ws.cell_value(header_row - 1, c)).strip() for c in range(ws.ncols) if ws.cell_type(header_row - 1, c) != 0]

    raise BadRequest(detail=f"不支持的格式: {ext}")


async def preview_headers(
    db: AsyncSession,
    mapping_id: uuid.UUID,
    file_data: bytes,
    file_name: str,
) -> PreviewResponse:
    """读取 Excel 表头，与指定模板做匹配预览。

    Args:
        db: 数据库会话。
        mapping_id: 映射模板 ID。
        file_data: 文件字节内容。
        file_name: 原始文件名（用于判断扩展名）。

    Returns:
        PreviewResponse — 每条表头的匹配结果。
    """
    mapping = await get_mapping(db, mapping_id)
    ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""

    if ext not in ("xlsx", "xls", "csv"):
        raise BadRequest(detail=f"不支持的文件格式: .{ext}")

    headers = _read_excel_headers(file_data, ext, mapping.sheet_index, mapping.header_row)
    if not headers:
        raise BadRequest(detail="未能读取到表头行，请检查 header_row 参数")

    rules: list[dict] = mapping.rules.get("rules", [])
    matched_count = 0

    results: list[HeaderMatchResult] = []
    for h in headers:
        match = next((r for r in rules if r["user_header"] == h), None)
        if match:
            results.append(HeaderMatchResult(
                user_header=h,
                matched=True,
                system_field=match["system_field"],
                converter=match.get("converter", ""),
            ))
            matched_count += 1
        else:
            results.append(HeaderMatchResult(user_header=h, matched=False))

    return PreviewResponse(
        mapping_id=mapping.id,
        mapping_name=mapping.mapping_name,
        total_headers=len(headers),
        matched=matched_count,
        unmatched=len(headers) - matched_count,
        headers=results,
    )
