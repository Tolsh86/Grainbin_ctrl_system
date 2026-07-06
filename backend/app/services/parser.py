"""Excel 解析 + 数据清洗归一化

供 Celery 任务调用，不在 HTTP 层暴露。
"""

from __future__ import annotations

import csv
import uuid
from datetime import date
from io import BytesIO, StringIO

from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from xlrd import open_workbook, xldate_as_tuple

from app.core.exceptions import BadRequest
from app.models.ingest import IngestBatch, IngestRow, IngestError, FieldMapping
from app.utils.converter import (
    apply_converter,
    chinese_date_to_iso,
    excel_serial_to_date,
    iso_date,
)
from app.utils.validator import DEFAULT_VALIDATION_RULES, ValidationResult, validate_row

BATCH_SIZE = 500


async def parse_and_normalize_batch(db: AsyncSession, batch_id: uuid.UUID) -> None:
    """解析批次 Excel，清洗归一化 + 校验后分批写入 IngestRow。

    完整流程（事务内原子执行）：
    1. 加载批次 + 映射模板
    2. 从 MinIO 读取文件
    3. 解析原始行（按格式分发）
    4. 合并单元格向下补全
    5. 清洗归一化（日期、金额转换等）
    6. 数据校验（5 类规则）→ is_valid / validation_flags
    7. 分批写入 IngestRow + IngestError
    8. 批次状态推进到 validated
    9. commit → 全部成功或全部回滚

    事务安全：所有 DB 写入在同一个 AsyncSession 事务内；任何异常
    导致整个事务回滚，不产生脏数据。
    """
    from app.utils.minio_client import minio_client as _minio

    # ── 1. 加载批次 ──
    batch = (await db.execute(
        select(IngestBatch).where(IngestBatch.id == batch_id).with_for_update()
    )).scalar_one_or_none()
    if not batch:
        raise ValueError(f"批次 {batch_id} 不存在")

    # 并发防护：解析中的批次不可重入
    if batch.status == "parsing":
        raise BadRequest(detail=f"批次 {batch_id} 正在解析中，请勿重复触发")

    # ── 2. 加载映射模板 ──
    if not batch.mapping_id:
        raise BadRequest(detail="批次缺少字段映射模板")
    mapping = (await db.execute(
        select(FieldMapping).where(FieldMapping.id == batch.mapping_id)
    )).scalar_one_or_none()
    if not mapping:
        raise BadRequest(detail=f"映射模板 {batch.mapping_id} 不存在")
    rules: list[dict] = mapping.rules.get("rules", [])
    if not rules:
        raise BadRequest(detail="映射模板规则为空")

    # ── 3. 从 MinIO 读取文件 ──
    try:
        _minio.client
    except RuntimeError:
        await _minio.init()
    file_data = _minio.download_bytes(batch.source_path)
    if not file_data:
        raise ValueError("MinIO 读取返回空数据")

    ext = batch.file_format or batch.source_doc.rsplit(".", 1)[-1].lower()

    # ── 4. 解析原始行 ──
    raw_rows = _read_excel_rows(file_data, ext, mapping.sheet_index, mapping.header_row, rules)
    if not raw_rows:
        batch.status = "validated"
        batch.total_rows = 0
        batch.parsed_rows = 0
        batch.valid_rows = 0
        batch.error_rows = 0
        batch.quality_score = 100.0
        await db.commit()
        logger.info(f"批次 {batch_id} 无数据行（空文件），状态设为 validated")
        return

    # ── 5. 合并单元格补全 ──
    _fill_merged_cells(raw_rows)

    # ── 6. 清洗归一化 ──
    normalized_rows = [_normalize_row(r, rules) for r in raw_rows]

    # ── 6.5 数据校验 ──
    all_validation_results: list[list[ValidationResult]] = []
    for i, norm in enumerate(normalized_rows):
        # 唯一性校验：当前行与其他所有 normalized 行对比
        other_rows = [n for j, n in enumerate(normalized_rows) if j != i]
        results = validate_row(norm, DEFAULT_VALIDATION_RULES, other_rows)
        all_validation_results.append(results)

    # ── 7. 分批写入 IngestRow + IngestError ──
    total = len(normalized_rows)
    valid_count = 0
    error_count = 0

    batch.total_rows = total
    batch.parsed_rows = total

    for i in range(0, total, BATCH_SIZE):
        chunk_raw = raw_rows[i : i + BATCH_SIZE]
        chunk_norm = normalized_rows[i : i + BATCH_SIZE]
        chunk_validations = all_validation_results[i : i + BATCH_SIZE]

        for j, (raw, norm, v_results) in enumerate(zip(chunk_raw, chunk_norm, chunk_validations)):
            row_no = i + j + 1

            # 构建 validation_flags
            flags: dict[str, dict] = {}
            for vr in v_results:
                key = f"{vr.field}__{vr.code}"
                flags[key] = {
                    "rule_type": vr.rule_type,
                    "severity": vr.severity,
                    "message": vr.message,
                }

            # 判定校验状态
            has_error = any(r.severity == "error" for r in v_results)
            has_warning = any(r.severity == "warning" for r in v_results)
            has_info = any(r.severity == "info" for r in v_results)

            if has_error:
                validation_status = "error"
                error_count += 1
            elif has_warning:
                validation_status = "warning"
                valid_count += 1
            elif has_info:
                validation_status = "suspicious"
                valid_count += 1
            else:
                validation_status = "normal"
                valid_count += 1

            row_is_valid = not has_error

            # 写入清洗行
            row = IngestRow(
                batch_id=batch.id,
                row_no=row_no,
                raw_payload=raw,
                normalized=norm,
                mapped={},
                project_id=batch.project_id,
                data_date=norm.get("data_date"),
                category=norm.get("category"),
                item_name=norm.get("item_name"),
                planned_quantity=norm.get("planned_quantity"),
                actual_quantity=norm.get("actual_quantity"),
                unit=norm.get("unit"),
                unit_price=norm.get("unit_price"),
                amount=norm.get("amount"),
                cost_type=norm.get("cost_type"),
                is_valid=row_is_valid,
                validation_flags=flags,
                validation_status=validation_status,
                quality_score=None if has_error else 100.0,
            )
            db.add(row)

            # 写入校验错误到 IngestError
            for vr in v_results:
                if vr.severity == "error":
                    error_rec = IngestError(
                        batch_id=batch.id,
                        row_id=row.id,
                        error_stage="validate",
                        error_code=vr.code,
                        error_message=vr.message,
                        error_field=vr.field,
                        error_value=str(norm.get(vr.field)) if vr.field in norm else None,
                        severity=vr.severity,
                    )
                    db.add(error_rec)

        await db.flush()

    # ── 8. 批次状态推进 → validated ──
    batch.valid_rows = valid_count
    batch.error_rows = error_count
    batch.status = "validated"

    # 质量分计算：校验通过率 × 100
    if total > 0:
        batch.quality_score = round(valid_count / total * 100, 2)

    # ── 9. 事务提交：全成功或全回滚 ──
    await db.commit()
    logger.info(
        f"批次 {batch_id} 解析+校验完成: {total} 行, "
        f"valid={valid_count}, error={error_count}, "
        f"quality={batch.quality_score}%"
    )


# ═══════════════════════════════════════════════════════
# 文件读取
# ═══════════════════════════════════════════════════════


def _build_col_map(headers: list[str], rules: list[dict]) -> dict[int, dict]:
    """根据表头行匹配映射规则，返回 {列索引: rule}。"""
    rule_map = {r["user_header"]: r for r in rules}
    col_map: dict[int, dict] = {}
    for idx, h in enumerate(headers):
        h_clean = str(h).strip() if h else ""
        if h_clean in rule_map:
            col_map[idx] = rule_map[h_clean]
    return col_map


def _read_excel_rows(
    file_data: bytes,
    ext: str,
    sheet_index: int,
    header_row: int,
    rules: list[dict],
) -> list[dict[str, object]]:
    if ext == "csv":
        return _read_csv_rows(file_data, header_row, rules)
    if ext == "xlsx":
        return _read_xlsx_rows(file_data, sheet_index, header_row, rules)
    if ext == "xls":
        return _read_xls_rows(file_data, sheet_index, header_row, rules)
    raise BadRequest(detail=f"不支持的文件格式: {ext}")


def _read_csv_rows(file_data: bytes, header_row: int, rules: list[dict]) -> list[dict[str, object]]:
    try:
        text = file_data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_data.decode("gbk")
    reader = csv.reader(StringIO(text))

    # 跳到表头行
    for _ in range(header_row - 1):
        next(reader, None)
    try:
        headers = next(reader)
    except StopIteration:
        return []
    col_map = _build_col_map(list(headers), rules)
    if not col_map:
        return []

    rows: list[dict[str, object]] = []
    for line in reader:
        raw: dict[str, object] = {}
        for idx, rule in col_map.items():
            raw[rule["user_header"]] = line[idx] if idx < len(line) else None
        rows.append(raw)
    return rows


def _read_xlsx_rows(
    file_data: bytes,
    sheet_index: int,
    header_row: int,
    rules: list[dict],
) -> list[dict[str, object]]:
    wb = load_workbook(BytesIO(file_data), read_only=True, data_only=True)
    try:
        if sheet_index >= len(wb.sheetnames):
            raise BadRequest(detail=f"Sheet 索引 {sheet_index} 超出范围（共 {len(wb.sheetnames)} 个）")
        ws = wb[wb.sheetnames[sheet_index]]

        # 读表头
        header_iter = ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        header_row_data = next(header_iter, [])
        headers = [str(c).strip() if c is not None else "" for c in header_row_data]
        col_map = _build_col_map(headers, rules)
        if not col_map:
            return []

        # 读数据行
        rows: list[dict[str, object]] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            raw: dict[str, object] = {}
            for idx, rule in col_map.items():
                raw[rule["user_header"]] = row[idx] if idx < len(row) else None
            rows.append(raw)
        return rows
    finally:
        wb.close()


def _read_xls_rows(
    file_data: bytes,
    sheet_index: int,
    header_row: int,
    rules: list[dict],
) -> list[dict[str, object]]:
    wb = open_workbook(file_contents=file_data)
    if sheet_index >= wb.nsheets:
        raise BadRequest(detail=f"Sheet 索引 {sheet_index} 超出范围（共 {wb.nsheets} 个）")
    ws = wb.sheet_by_index(sheet_index)

    # 读表头（header_row 是 1-based → xlrd 需要 0-based）
    if header_row > ws.nrows:
        raise BadRequest(detail=f"表头行 {header_row} 超出文件总行数 {ws.nrows}")
    headers = [str(ws.cell_value(header_row - 1, c)).strip() for c in range(ws.ncols)]
    col_map = _build_col_map(headers, rules)
    if not col_map:
        return []

    # 读数据行
    rows: list[dict[str, object]] = []
    for r in range(header_row, ws.nrows):
        raw: dict[str, object] = {}
        for idx, rule in col_map.items():
            if idx >= ws.ncols:
                raw[rule["user_header"]] = None
                continue
            cell_type = ws.cell_type(r, idx)
            cell_value = ws.cell_value(r, idx)
            if cell_type == 3:  # XL_CELL_DATE
                try:
                    dt = xldate_as_tuple(cell_value, wb.datemode)
                    raw[rule["user_header"]] = date(*dt[:3])
                except Exception:
                    raw[rule["user_header"]] = cell_value
            elif cell_type == 0:  # XL_CELL_EMPTY
                raw[rule["user_header"]] = None
            else:
                raw[rule["user_header"]] = cell_value
        rows.append(raw)
    return rows


# ═══════════════════════════════════════════════════════
# 合并单元格补全
# ═══════════════════════════════════════════════════════


def _fill_merged_cells(rows: list[dict[str, object]]) -> None:
    """合并单元格向下补全空值（原地修改）。

    当某行某字段为 None 或空串时，用上一行同字段的值填充。
    """
    if not rows:
        return
    keys = list(rows[0].keys())
    for i in range(1, len(rows)):
        for key in keys:
            val = rows[i].get(key)
            if val is None or val == "":
                rows[i][key] = rows[i - 1].get(key)


# ═══════════════════════════════════════════════════════
# 清洗归一化
# ═══════════════════════════════════════════════════════


def _clean_cell(value: object) -> object:
    """空格去除，空值 → None。"""
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return value


def _normalize_row(raw: dict[str, object], rules: list[dict]) -> dict[str, object]:
    """清洗一行：去空格、日期自动检测、转换器应用。"""
    normalized: dict[str, object] = {}
    for rule in rules:
        user_header = rule["user_header"]
        system_field = rule["system_field"]
        converter_name = rule.get("converter", "")

        value = _clean_cell(raw.get(user_header))

        # 日期自动检测：数值 → Excel 序列号，字符串 → ISO/中文日期
        if value is not None and (
            system_field == "data_date"
            or converter_name in ("excel_serial_to_date", "chinese_date_to_iso", "iso_date")
        ):
            if isinstance(value, (int, float)):
                date_val = excel_serial_to_date(value)
                if date_val:
                    value = date_val
            elif isinstance(value, str):
                date_val = iso_date(value) or chinese_date_to_iso(value)
                if date_val:
                    value = date_val

        # 显式转换器
        if converter_name and value is not None:
            value = apply_converter(value, converter_name)

        normalized[system_field] = value
    return normalized
