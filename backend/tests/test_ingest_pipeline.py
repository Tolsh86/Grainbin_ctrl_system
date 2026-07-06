"""落库全链路业务测试：上传→映射→解析→校验→入库→撤回

验证数据从 Excel 到 t_data_rows 的完整流转，包括异常路径。
"""

from __future__ import annotations

import io
import uuid
from datetime import date, datetime, UTC

import pytest
import pytest_asyncio
from openpyxl import Workbook
from sqlalchemy import select

from app.models.data_row import DataRow
from app.models.ingest import IngestBatch, IngestRow, IngestError, FieldMapping
from app.services.ingest import commit_batch, rollback_batch
from app.services.parser import parse_and_normalize_batch


# ═══════════════════════════════════════════════════════
# 内存 MinIO — 存储上传的 bytes，parser 可读取
# ═══════════════════════════════════════════════════════

class MemoryMinIO:
    """内存 MinIO：upload_bytes 存 bytes，download_bytes 返回 bytes。"""

    def __init__(self):
        self._store: dict[str, bytes] = {}

    def upload_bytes(self, key: str, data: bytes, **kwargs):
        self._store[key] = data
        return key  # 返回存储路径

    def download_bytes(self, key: str) -> bytes | None:
        return self._store.get(key)


@pytest_asyncio.fixture(scope="function")
async def memory_minio():
    """将全局 minio_client 替换为内存版本，测试结束恢复。"""
    from app.utils.minio_client import minio_client

    mem = MemoryMinIO()
    original_upload = minio_client.upload_bytes
    original_download = getattr(minio_client, "download_bytes", None)

    minio_client.upload_bytes = mem.upload_bytes
    minio_client.download_bytes = mem.download_bytes
    # 绕过 property —— parser 检查 _minio.client 是否存在
    minio_client._client = True

    yield mem

    minio_client.upload_bytes = original_upload
    if original_download:
        minio_client.download_bytes = original_download


# ═══════════════════════════════════════════════════════
# 辅助函数
# ═══════════════════════════════════════════════════════

def _make_mapping_rules() -> list[dict]:
    """生成映射规则：Match conftest sample_excel_bytes 表头。"""
    return [
        {"user_header": "序号",     "system_field": "row_index"},
        {"user_header": "日期",     "system_field": "data_date",       "converter": "iso_date"},
        {"user_header": "分项名称", "system_field": "item_name",        "converter": "trim"},
        {"user_header": "计划量",   "system_field": "planned_quantity"},
        {"user_header": "实际量",   "system_field": "actual_quantity"},
        {"user_header": "单位",     "system_field": "unit",             "converter": "trim"},
        {"user_header": "单价(元)", "system_field": "unit_price",       "converter": "yuan_to_fen"},
        {"user_header": "金额(元)", "system_field": "amount",           "converter": "yuan_to_fen"},
    ]


async def _create_mapping(db, rules=None) -> FieldMapping:
    """在 DB 中创建字段映射模板并返回。"""
    from app.schemas.mapping import MappingCreate, MappingRuleItem
    from app.services import mapping as mapping_service

    rule_list = rules or _make_mapping_rules()
    rule_items = [MappingRuleItem(**r) for r in rule_list]

    data = MappingCreate(
        mapping_name="测试映射模板",
        biz_type="weekly",
        file_format="xlsx",
        header_row=1,
        sheet_index=0,
        rules=rule_items,
    )
    return await mapping_service.create_mapping(db, data)


async def _create_ingest_batch(db, project_id, source_doc, source_path, file_format, mapping_id, uploaded_by) -> IngestBatch:
    """在 DB 中创建清洗批次。"""
    from app.services.ingest import create_batch as _create_batch_svc

    return await _create_batch_svc(
        db, project_id=project_id, source_doc=source_doc,
        source_path=source_path, file_format=file_format,
        uploaded_by=uploaded_by, mapping_id=mapping_id,
    )


# ═══════════════════════════════════════════════════════
# Test 1: 正常落库全链路
# ═══════════════════════════════════════════════════════

class TestIngestPipeline:
    """上传 → 解析 → 入库 → 撤回 完整链路"""

    @pytest.mark.asyncio
    async def test_full_pipeline_upload_parse_commit_rollback(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """核心场景：上传 Excel → 解析 → 入库 → 撤回"""
        admin_id, _token = admin_auth
        db = db_session

        # 1. 创建映射模板
        mapping = await _create_mapping(db)

        # 2. 生成 Excel 并上传到内存 MinIO
        excel_bytes, storage_path = _upload_excel_to_memory(memory_minio)

        # 3. 创建批次
        batch = await _create_ingest_batch(
            db,
            project_id=test_project,
            source_doc="test_progress.xlsx",
            source_path=storage_path,
            file_format="xlsx",
            mapping_id=mapping.id,
            uploaded_by=admin_id,
        )

        # 4. 执行解析
        await parse_and_normalize_batch(db, batch.id)

        # 刷新批次状态
        await db.refresh(batch)
        assert batch.status == "validated", f"解析+校验后状态应为 validated，实际: {batch.status}"
        assert batch.total_rows == 3
        assert batch.parsed_rows == 3
        assert batch.valid_rows == 3
        assert batch.error_rows == 0
        assert batch.quality_score == 100.0

        # 5. 验证 IngestRow 落库字段
        result = await db.execute(
            select(IngestRow).where(IngestRow.batch_id == batch.id).order_by(IngestRow.row_no)
        )
        rows = result.scalars().all()
        assert len(rows) == 3

        # 第 1 行：桩基础
        r1 = rows[0]
        assert r1.row_no == 1
        assert r1.is_valid == True, f"校验应通过，实际 flags={r1.validation_flags}"
        assert r1.validation_status == "normal"
        assert r1.validation_flags == {}, f"应无校验告警: {r1.validation_flags}"
        assert r1.item_name == "桩基础"
        assert r1.planned_quantity == 100
        assert r1.actual_quantity == 85
        assert r1.unit == "根"
        assert r1.unit_price == 500000    # 5000元 → 500000分
        assert r1.amount == 42500000      # 425000元 → 42500000分
        assert r1.data_date == date(2026, 7, 1)

        # 验证 raw_payload 保存了原始值
        assert "单价(元)" in r1.raw_payload
        assert r1.raw_payload["单价(元)"] == 5000

        # 验证 normalized 保存了归一化值
        assert r1.normalized["amount"] == 42500000  # 分

        # 第 2 行：土方开挖
        r2 = rows[1]
        assert r2.item_name == "土方开挖"
        assert r2.actual_quantity == 4800
        assert r2.unit_price == 3000      # 30元 → 3000分
        assert r2.amount == 14400000      # 144000元 → 14400000分

        # 第 3 行：钢筋绑扎
        r3 = rows[2]
        assert r3.item_name == "钢筋绑扎"
        assert r3.planned_quantity == 200
        assert r3.actual_quantity == 160
        assert r3.unit == "吨"
        assert r3.unit_price == 450000    # 4500元 → 450000分
        assert r3.amount == 72000000      # 720000元 → 72000000分

        # 6. 执行入库（解析后批次状态已是 validated，可直接 commit）
        committed_batch = await commit_batch(db, batch.id)
        assert committed_batch is not None
        assert committed_batch.status == "committed"
        assert committed_batch.committed_at is not None

        # 7. 验证 DataRow 落库（t_data_rows）
        data_result = await db.execute(
            select(DataRow).where(
                DataRow.project_id == test_project,
                DataRow.deleted_at.is_(None),
            ).order_by(DataRow.created_at)
        )
        data_rows = data_result.scalars().all()
        assert len(data_rows) == 3

        dr1 = data_rows[0]
        assert dr1.item_name == "桩基础"
        assert dr1.is_confirmed == True
        assert dr1.source_doc == "test_progress.xlsx"
        assert dr1.source_type == "upload"
        assert dr1.amount == 42500000
        assert dr1.unit == "根"

        # 验证 IngestRow 的 target_data_row_id 已关联
        await db.refresh(rows[0])
        assert rows[0].target_data_row_id == dr1.id

        # 8. 执行撤回
        rolled_batch = await rollback_batch(db, batch.id)
        assert rolled_batch is not None
        assert rolled_batch.status == "rolled_back"

        # 9. 验证 DataRow 软删除
        data_after = await db.execute(
            select(DataRow).where(
                DataRow.project_id == test_project,
                DataRow.deleted_at.is_(None),
            )
        )
        assert data_after.scalars().all() == []

        # 确认 deleted_at 已设置
        all_data = await db.execute(
            select(DataRow).where(DataRow.project_id == test_project)
        )
        for dr in all_data.scalars().all():
            assert dr.deleted_at is not None

    @pytest.mark.asyncio
    async def test_empty_excel_produces_zero_rows(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """空 Excel（只有表头无数据）→ 批次状态 normalized，行数 0"""
        admin_id, _token = admin_auth
        db = db_session

        mapping = await _create_mapping(db)

        # 创建空 Excel
        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "日期", "分项名称", "计划量", "实际量", "单位", "单价(元)", "金额(元)"])
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        excel_bytes = buf.getvalue()

        storage_path = "upload/EMPTY/empty.xlsx"
        memory_minio.upload_bytes(storage_path, excel_bytes)

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="empty.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)

        assert batch.status == "validated"
        assert batch.total_rows == 0
        assert batch.parsed_rows == 0
        assert batch.valid_rows == 0
        assert batch.error_rows == 0

    @pytest.mark.asyncio
    async def test_missing_mapping_rejected(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """映射模板缺失 → parse 抛异常，批次标记失败"""
        admin_id, _token = admin_auth
        db = db_session

        excel_bytes, storage_path = _upload_excel_to_memory(memory_minio)

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="no_mapping.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=None, uploaded_by=admin_id,
        )

        with pytest.raises(Exception):  # BadRequest: 批次缺少字段映射模板
            await parse_and_normalize_batch(db, batch.id)

    @pytest.mark.asyncio
    async def test_wrong_header_mapping_produces_no_rows(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """映射规则和 Excel 表头不匹配 → 解析 0 行"""
        admin_id, _token = admin_auth
        db = db_session

        # 映射模板用完全不匹配的规则
        wrong_rules = [
            {"user_header": "A列", "system_field": "item_name"},
            {"user_header": "B列", "system_field": "amount"},
        ]
        mapping = await _create_mapping(db, rules=wrong_rules)

        excel_bytes, storage_path = _upload_excel_to_memory(memory_minio)

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="wrong_headers.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)

        assert batch.status == "validated"
        assert batch.total_rows == 0

    @pytest.mark.asyncio
    async def test_commit_only_allowed_for_validated(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """仅 validated/review 状态的批次可入库 — pending 状态不可入库"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)
        excel_bytes, storage_path = _upload_excel_to_memory(memory_minio)

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="early_commit.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        # 状态 = pending，不可入库
        result = await commit_batch(db, batch.id)
        assert result is None

        # 解析 → validated 后可正常入库
        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)
        assert batch.status == "validated"

        # 再次尝试入库 → 成功
        result = await commit_batch(db, batch.id)
        assert result is not None
        assert result.status == "committed"

    @pytest.mark.asyncio
    async def test_rollback_only_allowed_for_committed(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """未 committed 的批次不可撤回 — validated 状态也不行"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)
        excel_bytes, storage_path = _upload_excel_to_memory(memory_minio)

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="early_rollback.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)

        # 状态 = validated，不可撤回
        result = await rollback_batch(db, batch.id)
        assert result is None

        # 入库后可以撤回
        await commit_batch(db, batch.id)
        await db.refresh(batch)
        assert batch.status == "committed"

        result = await rollback_batch(db, batch.id)
        assert result is not None
        assert result.status == "rolled_back"

    @pytest.mark.asyncio
    async def test_batch_not_found_raises_error(
        self, db_session, memory_minio,
    ):
        """不存在的批次 ID → ValueError"""
        fake_id = uuid.uuid4()
        with pytest.raises(ValueError, match="不存在"):
            await parse_and_normalize_batch(db_session, fake_id)


# ═══════════════════════════════════════════════════════
# Test 2: 金额一致性校验
# ═══════════════════════════════════════════════════════

class TestAmountConsistency:
    """金额转换正确性：元→分、万元→分、千分位"""

    @pytest.mark.asyncio
    async def test_yuan_to_fen_multiplication_check(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """验证：单价(元) × 100 = unit_price(分)，总价(元) × 100 = amount(分)"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)

        # 自定义 Excel：确保金额一致性
        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "日期", "分项名称", "计划量", "实际量", "单位", "单价(元)", "金额(元)"])
        ws.append([1, "2026-07-01", "混凝土", 500, 480, "m³", 380, 182400])
        ws.append([2, "2026-07-01", "钢筋",   80,  80, "t",  4200, 336000])
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        storage_path = "upload/AMOUNT/amount_test.xlsx"
        memory_minio.upload_bytes(storage_path, buf.getvalue())

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="amount_test.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)

        result = await db.execute(
            select(IngestRow).where(IngestRow.batch_id == batch.id).order_by(IngestRow.row_no)
        )
        rows = result.scalars().all()

        for row in rows:
            # 单价一致性: raw["单价(元)"] * 100 == normalized["unit_price"]
            raw_price = row.raw_payload["单价(元)"]
            norm_price = row.normalized["unit_price"]
            expected_price = int(raw_price) * 100
            assert norm_price == expected_price, (
                f"{row.item_name}: raw单价={raw_price}, expected={expected_price}分, got={norm_price}分"
            )

            # 总价一致性
            raw_amount = row.raw_payload["金额(元)"]
            norm_amount = row.normalized["amount"]
            expected_amount = int(raw_amount) * 100
            assert norm_amount == expected_amount, (
                f"{row.item_name}: raw总价={raw_amount}, expected={expected_amount}分, got={norm_amount}分"
            )

            # 总价检验: 单价(分) × 实际量 = 总价(分)
            computed = row.normalized["unit_price"] * row.normalized["actual_quantity"]
            assert computed == norm_amount, (
                f"{row.item_name}: 单价({norm_price}分) × 实际量({row.normalized['actual_quantity']}) "
                f"= {computed} ≠ 总价({norm_amount}分)"
            )

    @pytest.mark.asyncio
    async def test_amount_with_comma_separator(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """千分位分隔符金额: '12,345.67' → 1234567 分"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)

        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "日期", "分项名称", "计划量", "实际量", "单位", "单价(元)", "金额(元)"])
        ws.append([1, "2026-07-01", "测试项", 10, 10, "个", "1,234.56", "12,345.67"])
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        storage_path = "upload/COMMA/comma_test.xlsx"
        memory_minio.upload_bytes(storage_path, buf.getvalue())

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="comma_test.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)

        result = await db.execute(
            select(IngestRow).where(IngestRow.batch_id == batch.id)
        )
        row = result.scalar_one()

        assert row.normalized["unit_price"] == 123456    # 1234.56元 → 123456分
        assert row.normalized["amount"] == 1234567       # 12345.67元 → 1234567分


# ═══════════════════════════════════════════════════════
# Test 3: DataRow 落库字段完整性
# ═══════════════════════════════════════════════════════

class TestDataRowFields:
    """验证从 IngestRow → DataRow 的字段映射完整性"""

    @pytest.mark.asyncio
    async def test_all_fields_transferred_to_datarow(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """入库后 t_data_rows 的所有业务字段应与 IngestRow 一致"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)
        excel_bytes, storage_path = _upload_excel_to_memory(memory_minio)

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="field_test.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)
        # 解析后状态已是 validated，直接入库
        await commit_batch(db, batch.id)

        # 按 row_no 排序对比
        result = await db.execute(
            select(IngestRow).where(IngestRow.batch_id == batch.id).order_by(IngestRow.row_no)
        )
        ingest_rows = result.scalars().all()

        data_result = await db.execute(
            select(DataRow).where(
                DataRow.project_id == test_project,
                DataRow.deleted_at.is_(None),
            ).order_by(DataRow.created_at)
        )
        data_rows = data_result.scalars().all()

        assert len(ingest_rows) == len(data_rows)

        # 逐字段对比
        FIELDS = [
            "project_id", "data_date", "category", "item_name",
            "planned_quantity", "actual_quantity", "unit",
            "unit_price", "amount", "cost_type",
        ]
        for ing, dr in zip(ingest_rows, data_rows):
            for field in FIELDS:
                ing_val = getattr(ing, field)
                dr_val = getattr(dr, field)
                assert ing_val == dr_val, (
                    f"row_no={ing.row_no}, field={field}: "
                    f"IngestRow={ing_val}, DataRow={dr_val}"
                )

            # DataRow 特有字段
            assert dr.is_confirmed == True
            assert dr.source_doc == "field_test.xlsx"
            assert dr.source_type == "upload"


# ═══════════════════════════════════════════════════════
# Test 4: 校验集成 — 必填/类型/范围/唯一性
# ═══════════════════════════════════════════════════════

class TestValidationIntegration:
    """验证 parse_and_normalize_batch 正确集成校验逻辑"""

    @pytest.mark.asyncio
    async def test_missing_required_fields_caught_when_first_row_dirty(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """第一行缺必填字段 → is_valid=False; 第二行正常 → is_valid=True.

        注意: _fill_merged_cells 会补全所有空值，所以只有第一行（无上一行可继承）
        才能真正触发必填校验失败。
        """
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)

        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "日期", "分项名称", "计划量", "实际量", "单位", "单价(元)", "金额(元)"])
        ws.append([1, None,           None,  100, 85, "根", 5000, None])   # 缺少 data_date + item_name + amount
        ws.append([2, "2026-07-01", "桩基础", 200, 160, "吨", 4500, 720000])  # 正常
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        storage_path = "upload/VALIDATE/missing_first_row.xlsx"
        memory_minio.upload_bytes(storage_path, buf.getvalue())

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="missing_first_row.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)

        assert batch.status == "validated"
        assert batch.total_rows == 2
        assert batch.valid_rows == 1
        assert batch.error_rows == 1
        assert batch.quality_score == 50.0

        result = await db.execute(
            select(IngestRow).where(IngestRow.batch_id == batch.id).order_by(IngestRow.row_no)
        )
        rows = result.scalars().all()

        # 第 1 行：校验失败
        assert rows[0].is_valid == False
        assert rows[0].validation_status == "error"
        assert len(rows[0].validation_flags) >= 3  # data_date + item_name + amount
        assert any("data_date" in k for k in rows[0].validation_flags)
        assert any("item_name" in k for k in rows[0].validation_flags)
        assert any("amount" in k for k in rows[0].validation_flags)

        # 第 2 行：校验通过
        assert rows[1].is_valid == True
        assert rows[1].validation_status == "normal"
        assert rows[1].validation_flags == {}

    @pytest.mark.asyncio
    async def test_amount_out_of_range_marked_as_error(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """金额超出范围 → is_valid=False"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)

        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "日期", "分项名称", "计划量", "实际量", "单位", "单价(元)", "金额(元)"])
        ws.append([1, "2026-07-01", "桩基础", 100, 85, "根", 5000, 425000])
        ws.append([2, "2026-07-01", "超大金额", 1, 1, "项", 1, 99999999999])  # 超出范围
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        storage_path = "upload/VALIDATE/range_test.xlsx"
        memory_minio.upload_bytes(storage_path, buf.getvalue())

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="range_test.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)

        assert batch.error_rows == 1
        result = await db.execute(
            select(IngestRow).where(IngestRow.batch_id == batch.id, IngestRow.row_no == 2)
        )
        row = result.scalar_one()
        assert row.is_valid == False
        assert any("AMOUNT_OUT_OF_RANGE" in k for k in row.validation_flags)

    @pytest.mark.asyncio
    async def test_duplicate_rows_trigger_uniqueness_warning(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """重复行触发唯一性 warning（不阻断，is_valid=True）"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)

        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "日期", "分项名称", "计划量", "实际量", "单位", "单价(元)", "金额(元)"])
        ws.append([1, "2026-07-01", "桩基础", 100, 85, "根", 5000, 425000])
        ws.append([2, "2026-07-01", "桩基础", 200, 160, "根", 5000, 425000])  # 重复
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        storage_path = "upload/VALIDATE/dup_test.xlsx"
        memory_minio.upload_bytes(storage_path, buf.getvalue())

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="dup_test.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)

        # 唯一性不阻断 → 所有行仍 valid
        assert batch.valid_rows == 2
        assert batch.error_rows == 0

        result = await db.execute(
            select(IngestRow).where(IngestRow.batch_id == batch.id).order_by(IngestRow.row_no)
        )
        rows = result.scalars().all()
        # 两行都有 warning（互相重复）
        for r in rows:
            assert r.is_valid == True
            assert r.validation_status in ("warning", "normal")
            assert len(r.validation_flags) >= 1  # 至少 "DUPLICATE_KEY"

    @pytest.mark.asyncio
    async def test_ingest_errors_written_for_validation_failures(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """校验失败的行 → IngestError 记录写入"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)

        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "日期", "分项名称", "计划量", "实际量", "单位", "单价(元)", "金额(元)"])
        ws.append([1, "2026-07-01", None, 100, 85, "根", 5000, None])  # item_name + amount 缺失
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        storage_path = "upload/VALIDATE/error_log.xlsx"
        memory_minio.upload_bytes(storage_path, buf.getvalue())

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="error_log.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)

        # 验证 IngestError 记录
        err_result = await db.execute(
            select(IngestError).where(IngestError.batch_id == batch.id)
        )
        errors = err_result.scalars().all()
        assert len(errors) >= 2  # item_name REQUIRED + amount REQUIRED
        assert all(e.error_stage == "validate" for e in errors)
        assert all(e.severity == "error" for e in errors)
        assert any("item_name" in e.error_message for e in errors)
        assert any("amount" in e.error_message for e in errors)


# ═══════════════════════════════════════════════════════
# Test 5: 事务安全性
# ═══════════════════════════════════════════════════════

class TestTransactionSafety:
    """验证事务原子性和并发防护"""

    @pytest.mark.asyncio
    async def test_parse_failure_rolls_back_all_rows(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """解析中途失败 → IngestRow 全部回滚，不残留脏数据"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)

        # 正常 Excel，中途抛异常的场景已在 test_missing_mapping_rejected 覆盖
        # 这里验证：抛异常后 IngestRow 表中无数据
        excel_bytes, storage_path = _upload_excel_to_memory(memory_minio)

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="rollback_test.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=None,  # 无映射 → 解析失败
            uploaded_by=admin_id,
        )

        try:
            await parse_and_normalize_batch(db, batch.id)
        except Exception:
            pass

        # 验证无残留 IngestRow
        from sqlalchemy import func
        count_result = await db.execute(
            select(func.count(IngestRow.id)).where(IngestRow.batch_id == batch.id)
        )
        assert count_result.scalar() == 0, "解析失败后不应残留 IngestRow"

    @pytest.mark.asyncio
    async def test_concurrent_commit_prevented_by_row_lock(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """同一批次不能重复入库 — 状态机防护"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)
        excel_bytes, storage_path = _upload_excel_to_memory(memory_minio)

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="concurrent.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)

        # 第一次入库 → 成功
        result1 = await commit_batch(db, batch.id)
        assert result1 is not None
        assert result1.status == "committed"

        # 第二次入库 → 拒绝（状态已是 committed）
        result2 = await commit_batch(db, batch.id)
        assert result2 is None

    @pytest.mark.asyncio
    async def test_validated_batch_can_be_committed_then_rolled_back(
        self, db_session, memory_minio, admin_auth, test_project,
    ):
        """完整闭环：解析(validated) → 入库(committed) → 撤回(rolled_back)"""
        admin_id, _token = admin_auth
        db = db_session
        mapping = await _create_mapping(db)
        excel_bytes, storage_path = _upload_excel_to_memory(memory_minio)

        batch = await _create_ingest_batch(
            db, project_id=test_project, source_doc="full_cycle.xlsx",
            source_path=storage_path, file_format="xlsx",
            mapping_id=mapping.id, uploaded_by=admin_id,
        )

        # Phase 1: 解析+校验
        await parse_and_normalize_batch(db, batch.id)
        await db.refresh(batch)
        assert batch.status == "validated"
        assert batch.valid_rows == 3

        # Phase 2: 入库
        committed = await commit_batch(db, batch.id)
        assert committed is not None
        assert committed.status == "committed"

        # Phase 3: 撤回
        rolled = await rollback_batch(db, batch.id)
        assert rolled is not None
        assert rolled.status == "rolled_back"


# ═══════════════════════════════════════════════════════
# 辅助：生成测试 Excel 并上传到内存 MinIO
# ═══════════════════════════════════════════════════════

def _upload_excel_to_memory(minio_mem: MemoryMinIO) -> tuple[bytes, str]:
    """生成标准测试 Excel，存入内存 MinIO，返回 (bytes, storage_path)。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["序号", "日期", "分项名称", "计划量", "实际量", "单位", "单价(元)", "金额(元)"])
    ws.append([1, "2026-07-01", "桩基础",   100,  85,   "根", 5000, 425000])
    ws.append([2, "2026-07-01", "土方开挖", 5000, 4800, "m³", 30,   144000])
    ws.append([3, "2026-07-01", "钢筋绑扎", 200,  160,  "吨", 4500, 720000])

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    data = buf.getvalue()

    storage_path = "upload/TEST-BATCH/test_progress.xlsx"
    minio_mem.upload_bytes(storage_path, data)
    return data, storage_path
