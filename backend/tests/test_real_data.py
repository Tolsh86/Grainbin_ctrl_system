"""真实甲方数据格式测试：模拟粮仓过控系统 Excel 格式

基于「粮仓过控系统-表格」中的7类真实 Excel 文件构造测试数据，
验证 parser 在真实业务场景下的表现。
"""

from __future__ import annotations

import io
from datetime import date

import pytest
from openpyxl import Workbook

from app.services.parser import (
    _build_col_map,
    _clean_cell,
    _fill_merged_cells,
    _normalize_row,
    _read_xlsx_rows,
)


# ═══════════════════════════════════════════════════════
# Excel 构造工具：模拟甲方真实表格结构
# ═══════════════════════════════════════════════════════

WEEKLY_FIELDS = [
    "项目总投资",
    "2026年度计划投资",
    "2026年度已完成投资",
    "其中：建安工程费",
    "其中：二类费用",
    "本周完成建安投资",
    "本周二类费用完成金额",
    "上周完成建安投资",
    "上周二类费用完成金额",
]


def _make_weekly_percentage_xlsx(weeks: list[dict]) -> bytes:
    """「周报百分比计算表格」：每 sheet 一周，垂直键值对"""
    wb = Workbook()
    wb.remove(wb.active)
    for wk in weeks:
        ws = wb.create_sheet(title=wk["sheet_name"])
        for i, (k, v) in enumerate(wk["data"].items(), start=2):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=3, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_payment_review_xlsx(entries: list[dict]) -> bytes:
    """「进度款审核汇总表」：封面 + 各参建单位 sheet"""
    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = "封面"
    ws_cover.cell(row=2, column=2, value="川西现代粮食物流中心建设项目设计-采购-施工(cpc)")
    ws_cover.cell(row=4, column=2, value="进 度 款 审 核 表")
    ws_cover.cell(row=10, column=2, value="建 设 单 位：   四川德阳省食油储备库有限公司")

    for sheet_info in entries:
        ws = wb.create_sheet(title=sheet_info["sheet_name"])
        ws.cell(row=1, column=1, value=sheet_info["title"])
        ws.cell(row=2, column=1, value="工程名称：川西现代粮食物流中心建设项目设计-采购-施工(cpc)")
        ws.cell(row=3, column=1, value=f"申请单位：{sheet_info['applicant']}")
        ws.cell(row=3, column=9, value="单位：元")
        for j, h in enumerate(sheet_info["headers"], start=1):
            ws.cell(row=4, column=j, value=h)
        for i, row_data in enumerate(sheet_info["rows"], start=5):
            for j, val in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_monthly_report_xlsx(projects: list[dict]) -> bytes:
    """「月报（对产投）」：多级标题 + 宽表"""
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "新投项目"

    ws_main.cell(row=1, column=1, value="德阳发展集团2026年投资计划表（新投项目）")
    ws_main.cell(row=2, column=1, value="单位：万元")

    headers = [
        "序号", "项目名称", "项目性质", "投资时序", "业主单位",
        "项目内容（建设内容）", "实施主体", "投资性质", "投资结构",
        "投资领域", "实施时间", "计划总投资", "2026年计划投资",
        "本月完成投资额", "本年完成投资额", "累计完成投资额",
        "本月项目进展", "投资预期收益",
    ]
    for j, h in enumerate(headers, start=1):
        ws_main.cell(row=4, column=j, value=h)

    field_keys = [
        "序号", "项目名称", "项目性质", "投资时序", "业主单位",
        "项目内容（建设内容）", "实施主体", "投资性质", "投资结构",
        "投资领域", "实施时间", "计划总投资", "2026年计划投资",
        "本月完成投资额", "本年完成投资额", "累计完成投资额",
        "本月项目进展", "投资预期收益",
    ]
    for i, proj in enumerate(projects, start=5):
        for j, key in enumerate(field_keys, start=1):
            ws_main.cell(row=i, column=j, value=proj.get(key, None))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_progress_drc_xlsx(projects: list[dict]) -> bytes:
    """「发改月报」：省重点项目进度表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "省重"

    ws.cell(row=1, column=1, value="2026年德阳市省重点项目1-6月进度（续建+新开工）")

    headers = [
        "序号", "项目名称", "建设起止年限", "建设内容及规模",
        "计划总投资（万元）", "截至2025年底累计完成投资（万元）",
        "2026年预计投资（万元）", "本月完成投资",
        "本年完成投资", "完成率", "上月进展", "本月进展",
        "业主单位", "责任单位", "备注",
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=3, column=j, value=h)

    for i, proj in enumerate(projects, start=4):
        for j, key in enumerate(headers, start=1):
            ws.cell(row=i, column=j, value=proj.get(key, None))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_cost_ledger_xlsx(contracts: list[dict]) -> bytes:
    """「二类费用台账」：合同支付情况 + 多级表头"""
    wb = Workbook()
    ws = wb.active
    ws.title = "合同支付情况"

    ws.cell(row=1, column=1, value="合同支付情况")
    ws.cell(row=2, column=1, value="单位：元")

    base_headers = ["序号", "编号", "供应商", "合同名称", "签署日期", "合同金额"]
    pay_phase_headers = ["施工前阶段", "施工中阶段", "竣工阶段", "缺陷责任期"]
    summary_headers = ["累计已付（元）", "剩余未付（元）", "已付比例", "剩余未付比例"]
    detail_dates = [45962, 45992, 46023, 46054, 46082, 46113, 46143, 46174]

    for j, h in enumerate(base_headers, start=1):
        ws.cell(row=3, column=j, value=h)
    col_offset = len(base_headers) + 1
    ws.cell(row=3, column=col_offset, value="支付节点")
    for j, h in enumerate(summary_headers, start=col_offset + 5):
        ws.cell(row=3, column=j, value=h)
    pay_detail_start = col_offset + 5 + len(summary_headers) + 1
    ws.cell(row=3, column=pay_detail_start, value="支付明细")

    for j, h in enumerate(pay_phase_headers, start=col_offset + 1):
        ws.cell(row=4, column=j, value=h)
    for j, dt in enumerate(detail_dates, start=pay_detail_start):
        ws.cell(row=4, column=j, value=dt)

    field_keys = ["序号", "编号", "供应商", "合同名称", "签署日期", "合同金额"]
    for i, contract in enumerate(contracts, start=4):
        for j, key in enumerate(field_keys, start=1):
            ws.cell(row=i, column=j, value=contract.get(key, None))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════
# 测试 1：周报数据 — 垂直键值对（当前 parser 的局限性）
# ═══════════════════════════════════════════════════════

class TestWeeklyPercentage:
    """周报百分比计算表格：垂直键值对格式"""

    def test_vertical_format_not_supported_by_column_mapper(self):
        """垂直键值对格式（周报百分比表）：当前列映射 parser 无法处理

        原因：表头行全是空列名（A/B/C），_build_col_map 匹配不到任何规则。
        这是已知局限——垂直表需要特殊的 row-to-column 转换。
        已作为后续优化需求记录。
        """
        weeks_data = [{"sheet_name": "7.1", "data": {
            "项目总投资": 19320.33,
            "本周完成建安投资": 93.5,
        }}]
        xlsx_bytes = _make_weekly_percentage_xlsx(weeks_data)
        rules = _make_weekly_rules()
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=1, rules=rules)
        # 列映射表为空 → 返回空（垂直格式需要额外处理层）
        assert raw_rows == []

    def test_vertical_format_data_accessible_via_raw_read(self):
        """垂直表数据可以通过 openpyxl 直接读取（绕过列映射）"""
        from openpyxl import load_workbook
        weeks_data = [{"sheet_name": "7.1", "data": {
            "项目总投资": 19320.33,
            "本周完成建安投资": 93.5,
        }}]
        xlsx_bytes = _make_weekly_percentage_xlsx(weeks_data)
        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        assert len(rows) == 2
        assert rows[0][0] == "项目总投资"
        assert rows[0][2] == 19320.33


def _make_weekly_rules():
    return [
        {"user_header": "项目总投资", "system_field": "total_investment"},
        {"user_header": "2026年度计划投资", "system_field": "annual_plan"},
        {"user_header": "2026年度已完成投资", "system_field": "completed_ytd"},
        {"user_header": "其中：建安工程费", "system_field": "construction_cost"},
        {"user_header": "其中：二类费用", "system_field": "secondary_cost"},
        {"user_header": "本周完成建安投资", "system_field": "weekly_construction"},
        {"user_header": "本周二类费用完成金额", "system_field": "weekly_secondary"},
        {"user_header": "上周完成建安投资", "system_field": "last_week_construction"},
        {"user_header": "上周二类费用完成金额", "system_field": "last_week_secondary"},
    ]


# ═══════════════════════════════════════════════════════
# 测试 2：进度款审核 — 多 sheet 宽表
# ═══════════════════════════════════════════════════════

class TestPaymentReview:
    """进度款审核汇总表"""

    PAYMENT_RULES = [
        {"user_header": "序号", "system_field": "seq"},
        {"user_header": "期数/月份", "system_field": "period"},
        {"user_header": "累计报送产值", "system_field": "reported_output"},
        {"user_header": "合同金额", "system_field": "contract_amount"},
        {"user_header": "本期审核产值", "system_field": "reviewed_amount"},
        {"user_header": "累计审核产值", "system_field": "cumulative_reviewed"},
        {"user_header": "累计审核应付款", "system_field": "cumulative_payable"},
        {"user_header": "付款比例", "system_field": "payment_ratio"},
    ]

    @pytest.fixture
    def payment_xlsx(self):
        return _make_payment_review_xlsx([
            {
                "sheet_name": "施工单位进度产值审核汇总表",
                "title": "施工单位进度产值审核汇总表",
                "applicant": "德阳建设工程集团有限公司",
                "headers": [
                    "序号", "期数/月份", "累计报送产值", "合同金额",
                    "建筑安装工程费", "安全文明施工费",
                    "建筑安装工程费付款比例", "安全文明施工费付款比例",
                    "本期审核产值", "累计审核产值", "累计审核应付款", "备注",
                ],
                "rows": [
                    [1, "第一期/2026年01月", 10526301.85, 131729850,
                     97228260, 5356505.64, 0.1, 0.5,
                     10526301.85, 10526301.85, 10526301.85,
                     "分别汇入项目基本账户和农民工工资专用帐户"],
                    [2, "合计", 10526301.85, None, None, None, None, None,
                     10526301.85, 10526301.85, 10526301.85, None],
                ],
                "note": "注：以上金额均为含税金额，税率为9%。",
            },
            {
                "sheet_name": "设备单位进度产值审核汇总表",
                "title": "设备单位进度产值审核汇总表",
                "applicant": "中国机械工业建设集团有限公司",
                "headers": [
                    "序号", "期数/月份", "累计报送产值", "合同金额",
                    "付款比例", "本期审核产值", "累计审核产值", "累计审核应付款", "备注",
                ],
                "rows": [
                    [1, "第一期/2026年01月", 6512886, 32564430,
                     0.2, 6512886, 6512886, 6512886, None],
                ],
                "note": "注：以上金额均为含税金额，税率为9%。",
            },
        ])

    def test_construction_sheet_read(self, payment_xlsx):
        """施工单位 sheet — 表头在第4行，数据从第5行开始"""
        raw_rows = _read_xlsx_rows(payment_xlsx, sheet_index=1, header_row=4,
                                    rules=self.PAYMENT_RULES)
        assert len(raw_rows) == 2
        # keys 是 user_header（Excel 列名）
        assert "本期审核产值" in raw_rows[0]
        assert raw_rows[0]["本期审核产值"] == 10526301.85
        assert raw_rows[0]["期数/月份"] == "第一期/2026年01月"

    def test_equipment_sheet_read(self, payment_xlsx):
        """设备单位 sheet — 不同列结构独立读取"""
        raw_rows = _read_xlsx_rows(payment_xlsx, sheet_index=2, header_row=4,
                                    rules=self.PAYMENT_RULES)
        assert len(raw_rows) == 1
        assert raw_rows[0]["合同金额"] == 32564430
        assert raw_rows[0]["付款比例"] == 0.2

    def test_normalize_to_system_field(self, payment_xlsx):
        """归一化后映射到 system_field"""
        raw_rows = _read_xlsx_rows(payment_xlsx, sheet_index=1, header_row=4,
                                    rules=self.PAYMENT_RULES)
        result = _normalize_row(raw_rows[0], self.PAYMENT_RULES)

        # system_field 成为 key
        assert result["period"] == "第一期/2026年01月"
        assert result["reviewed_amount"] == 10526301.85
        assert result["contract_amount"] == 131729850

    def test_summary_row_contains_text(self, payment_xlsx):
        """合计行的期数列包含文本（不是数字）"""
        raw_rows = _read_xlsx_rows(payment_xlsx, sheet_index=1, header_row=4,
                                    rules=self.PAYMENT_RULES)
        result = _normalize_row(raw_rows[1], self.PAYMENT_RULES)
        assert result["period"] == "合计"

    def test_amount_in_yuan_conversion(self, payment_xlsx):
        """金额元转分 — 10526301.85 元 → 1052630185 分"""
        amt_rules = [
            {"user_header": "本期审核产值", "system_field": "amount",
             "converter": "yuan_to_fen"},
        ]
        raw_rows = _read_xlsx_rows(payment_xlsx, sheet_index=1, header_row=4,
                                    rules=amt_rules)
        result = _normalize_row(raw_rows[0], amt_rules)
        assert result["amount"] == 1052630185

    def test_cover_sheet_not_data(self, payment_xlsx):
        """封面 sheet 无匹配表头 → 返回空（不作为数据源）"""
        raw_rows = _read_xlsx_rows(payment_xlsx, sheet_index=0, header_row=4,
                                    rules=self.PAYMENT_RULES)
        # 封面 sheet 没有 "期数/月份" 等列名 → col_map 为空 → 返回 []
        assert raw_rows == []


# ═══════════════════════════════════════════════════════
# 测试 3：月报 — 多级表头 + 长文本 + 万元单位
# ═══════════════════════════════════════════════════════

class TestMonthlyReport:
    """月报（对产投）"""

    MONTHLY_RULES = [
        {"user_header": "项目名称", "system_field": "project_name"},
        {"user_header": "计划总投资", "system_field": "total_investment"},
        {"user_header": "2026年计划投资", "system_field": "annual_plan"},
        {"user_header": "本月完成投资额", "system_field": "monthly_investment"},
        {"user_header": "本年完成投资额", "system_field": "ytd_investment"},
        {"user_header": "累计完成投资额", "system_field": "cumulative_investment"},
        {"user_header": "项目内容（建设内容）", "system_field": "content"},
        {"user_header": "本月项目进展", "system_field": "progress"},
        {"user_header": "业主单位", "system_field": "owner"},
    ]

    def test_basic_read(self):
        """表头在第4行，数据从第5行开始"""
        projects = [{
            "项目名称": "川西现代粮食物流中心建设项目",
            "计划总投资": 19320.33,
            "2026年计划投资": 7500,
            "本月完成投资额": 1000,
        }]
        xlsx_bytes = _make_monthly_report_xlsx(projects)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=4,
                                    rules=self.MONTHLY_RULES)

        assert len(raw_rows) == 1
        # 原始 key 是 user_header
        assert raw_rows[0]["项目名称"] == "川西现代粮食物流中心建设项目"
        assert raw_rows[0]["计划总投资"] == 19320.33

    def test_normalize(self):
        """归一化 → system_field 为 key"""
        projects = [{"项目名称": "川西粮仓", "计划总投资": 19320.33}]
        xlsx_bytes = _make_monthly_report_xlsx(projects)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=4,
                                    rules=self.MONTHLY_RULES)
        result = _normalize_row(raw_rows[0], self.MONTHLY_RULES)
        assert result["project_name"] == "川西粮仓"
        assert result["total_investment"] == 19320.33

    def test_long_text(self):
        """建设内容包含长文本 → 原样保留"""
        long_content = (
            "新建7.25万吨高标准粮仓，其中浅圆仓5.25万吨，低温平房仓2万吨，"
            "配套建设卸粮棚、工作塔、仓间罩棚；新建4600吨食用油罐，"
            "配套建设油泵房、发油棚；新建质检楼，配套建设附属总平工程；"
            "仓顶阳光及附属设施。"
        )
        projects = [{"项目名称": "川西粮仓", "项目内容（建设内容）": long_content}]
        xlsx_bytes = _make_monthly_report_xlsx(projects)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=4,
                                    rules=self.MONTHLY_RULES)
        result = _normalize_row(raw_rows[0], self.MONTHLY_RULES)
        assert result["content"] == long_content

    def test_progress_with_newlines(self):
        """进展描述包含换行符 → 保留"""
        progress = "1.砖砌施工467个\n2.质检楼基坑开挖及支护完成\n3.场内临时道路完成约60%"
        projects = [{"项目名称": "川西粮仓", "本月项目进展": progress}]
        xlsx_bytes = _make_monthly_report_xlsx(projects)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=4,
                                    rules=self.MONTHLY_RULES)
        result = _normalize_row(raw_rows[0], self.MONTHLY_RULES)
        assert "\n" in result["progress"]
        assert "砖砌施工467个" in result["progress"]

    def test_null_optional_fields(self):
        """可选的数值字段为空 → None"""
        projects = [{"项目名称": "测试项目", "本月完成投资额": None, "计划总投资": 10000}]
        xlsx_bytes = _make_monthly_report_xlsx(projects)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=4,
                                    rules=self.MONTHLY_RULES)
        result = _normalize_row(raw_rows[0], self.MONTHLY_RULES)
        assert result["monthly_investment"] is None
        assert result["total_investment"] == 10000

    def test_multiple_projects(self):
        """批量多项目"""
        projects = [
            {"项目名称": "项目A", "计划总投资": 10000},
            {"项目名称": "项目B", "计划总投资": 20000},
            {"项目名称": "项目C", "计划总投资": 30000},
        ]
        xlsx_bytes = _make_monthly_report_xlsx(projects)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=4,
                                    rules=self.MONTHLY_RULES)
        assert len(raw_rows) == 3
        assert raw_rows[2]["项目名称"] == "项目C"
        assert raw_rows[2]["计划总投资"] == 30000


# ═══════════════════════════════════════════════════════
# 测试 4：发改月报 — 进度对比
# ═══════════════════════════════════════════════════════

class TestDRCMonthly:
    """发改月报：省市重点项目进度表"""

    DRC_RULES = [
        {"user_header": "项目名称", "system_field": "name"},
        {"user_header": "计划总投资（万元）", "system_field": "total_invest"},
        {"user_header": "本月完成投资", "system_field": "monthly"},
        {"user_header": "本年完成投资", "system_field": "ytd"},
        {"user_header": "完成率", "system_field": "completion_rate"},
        {"user_header": "本月进展", "system_field": "progress"},
        {"user_header": "建设起止年限", "system_field": "duration"},
        {"user_header": "业主单位", "system_field": "owner"},
    ]

    def test_basic_read(self):
        """表头在第3行，数据从第4行开始"""
        projects = [{
            "项目名称": "川西现代粮食物流中心建设项目",
            "建设起止年限": "2025-2027",
            "计划总投资（万元）": 19320.33,
            "本月完成投资": 1100,
            "本年完成投资": 2610,
            "完成率": 0.348,
        }]
        xlsx_bytes = _make_progress_drc_xlsx(projects)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=3,
                                    rules=self.DRC_RULES)

        assert len(raw_rows) == 1
        assert raw_rows[0]["项目名称"] == "川西现代粮食物流中心建设项目"

    def test_progress_comparison(self):
        """进度对比：本月 vs 本年累计 vs 年度计划"""
        projects = [{
            "项目名称": "川西粮仓",
            "计划总投资（万元）": 19320.33,
            "本月完成投资": 1100,
            "本年完成投资": 2610,
            "完成率": 0.348,
        }]
        xlsx_bytes = _make_progress_drc_xlsx(projects)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=3,
                                    rules=self.DRC_RULES)
        result = _normalize_row(raw_rows[0], self.DRC_RULES)

        # 三个时间维度
        assert result["monthly"] == 1100
        assert result["ytd"] == 2610
        assert result["total_invest"] == 19320.33
        # 完成率 = 2610 / 7500 = 不验证具体值，只验证被读到了
        assert result["completion_rate"] is not None

    def test_long_progress_description(self):
        """发改月报的进展描述特别长 → 完整保留"""
        long_progress = (
            "1.5#和6#平房仓基础已完成100%\n"
            "2.6#平房仓基础垫层浇筑完成90%\n"
            "3.质检楼基坑开挖及支护完成100%\n"
            "4.砖砌施工467个\n"
            "5.场内临时道路、场地硬化等完成约60%"
        )
        projects = [{
            "项目名称": "川西粮仓",
            "本月进展": long_progress,
        }]
        xlsx_bytes = _make_progress_drc_xlsx(projects)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=3,
                                    rules=self.DRC_RULES)
        result = _normalize_row(raw_rows[0], self.DRC_RULES)
        # 包含所有关键内容
        assert "5#和6#平房仓" in result["progress"]
        assert "砖砌施工467个" in result["progress"]


# ═══════════════════════════════════════════════════════
# 测试 5：二类费用台账 — 合同 + 支付节点
# ═══════════════════════════════════════════════════════

class TestCostLedger:
    """二类费用台账：合同支付情况"""

    CONTRACT_RULES = [
        {"user_header": "编号", "system_field": "contract_no"},
        {"user_header": "供应商", "system_field": "supplier"},
        {"user_header": "合同名称", "system_field": "contract_name"},
        {"user_header": "签署日期", "system_field": "sign_date", "converter": "iso_date"},
        {"user_header": "合同金额", "system_field": "contract_amount"},
    ]

    def test_basic_contract_read(self):
        """合同基本信息读取（header_row=3，数据从 row 4 开始）"""
        contracts = [{
            "编号": "CX001",
            "供应商": "四川兴恒通工程咨询有限公司",
            "合同名称": "四川省国家工程投资建设委托招标代理合同",
            "签署日期": date(2024, 8, 19),
            "合同金额": 50000,
        }]
        xlsx_bytes = _make_cost_ledger_xlsx(contracts)
        # header_row=3: 主列名在 row 3
        # 数据从 row 4 开始（不再有子表头行）
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=3,
                                    rules=self.CONTRACT_RULES)

        assert len(raw_rows) == 1
        assert raw_rows[0]["编号"] == "CX001"
        assert raw_rows[0]["合同金额"] == 50000

    def test_normalize_to_system_fields(self):
        """归一化后 key 为 system_field"""
        contracts = [{
            "编号": "CX016",
            "供应商": "四川兴天宇建设工程质量检测有限公司",
            "合同名称": "工程检测合同",
            "签署日期": date(2026, 2, 3),
            "合同金额": 120000,
        }]
        xlsx_bytes = _make_cost_ledger_xlsx(contracts)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=3,
                                    rules=self.CONTRACT_RULES)
        result = _normalize_row(raw_rows[0], self.CONTRACT_RULES)

        assert result["contract_no"] == "CX016"
        assert result["supplier"] == "四川兴天宇建设工程质量检测有限公司"
        assert result["contract_amount"] == 120000

    def test_null_contract_amount(self):
        """无合同金额（部分二类费用合同金额待定）"""
        contracts = [{
            "编号": "CX003",
            "供应商": "供应商C",
            "合同名称": "无金额合同",
            "合同金额": None,
        }]
        xlsx_bytes = _make_cost_ledger_xlsx(contracts)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=3,
                                    rules=self.CONTRACT_RULES)
        result = _normalize_row(raw_rows[0], self.CONTRACT_RULES)
        assert result["contract_amount"] is None

    def test_multiple_contracts(self):
        """多份合同批量读取"""
        contracts = [
            {"编号": "CX008", "供应商": "德阳建设工程集团有限公司",
             "合同名称": "建设安装工程", "合同金额": 131729850},
            {"编号": "CX009", "供应商": "中凯俊成建设咨询有限公司",
             "合同名称": "全过程咨询（监理）", "合同金额": 3040740},
            {"编号": "CX010", "供应商": "德阳兴源生态环境咨询有限公司",
             "合同名称": "环评咨询合同", "合同金额": 80000},
        ]
        xlsx_bytes = _make_cost_ledger_xlsx(contracts)
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=3,
                                    rules=self.CONTRACT_RULES)
        assert len(raw_rows) == 3
        # 建筑安装工程的金额
        assert raw_rows[0]["合同金额"] == 131729850
        # CX009
        assert raw_rows[1]["编号"] == "CX009"


# ═══════════════════════════════════════════════════════
# 测试 6：合并单元格 — 真实场景（多级分组）
# ═══════════════════════════════════════════════════════

class TestMergedCells:
    """模拟甲方报表中大量合并单元格的场景"""

    def test_two_level_project_subproject(self):
        """项目名合并 → 多行分项共享项目名"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        headers = ["项目名称", "分项工程", "实际产值（万元）", "审核产值（万元）"]
        for j, h in enumerate(headers, start=1):
            ws.cell(row=1, column=j, value=h)

        data = [
            ["川西粮食物流中心", "浅圆仓桩基", 500, 480],
            [None,               "平房仓基础", 300, 295],
            [None,               "质检楼主体", 200, 198],
            [None,               "油罐区",     150, 145],
            ["德阳南站枢纽",     "道路工程",   1000, 980],
            [None,               "地下停车场", 800, 795],
        ]
        for i, row_data in enumerate(data, start=2):
            for j, val in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        rules = [
            {"user_header": "项目名称", "system_field": "project"},
            {"user_header": "分项工程", "system_field": "sub_project"},
        ]
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=1, rules=rules)

        # 补全前：第2行项目名为 None
        assert raw_rows[1]["项目名称"] is None

        _fill_merged_cells(raw_rows)

        # 补全后：前4行 → 川西粮食物流中心
        for i in range(4):
            assert raw_rows[i]["项目名称"] == "川西粮食物流中心"
        # 后2行 → 德阳南站枢纽
        for i in range(4, 6):
            assert raw_rows[i]["项目名称"] == "德阳南站枢纽"

        # 分项工程列保持独立（不会被补全）
        assert raw_rows[0]["分项工程"] == "浅圆仓桩基"
        assert raw_rows[3]["分项工程"] == "油罐区"

    def test_three_level_hierarchy(self):
        """单位工程 → 分部工程 → 分项工程 三级合并"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        headers = ["单位工程", "分部工程", "分项工程", "产值（万元）"]
        for j, h in enumerate(headers, start=1):
            ws.cell(row=1, column=j, value=h)

        data = [
            ["粮仓主体", "基础工程", "桩基施工", 300],
            [None,       None,       "承台施工", 200],
            [None,       "主体结构", "墙体施工", 500],
            [None,       None,       "屋面施工", 180],
            ["附属工程", "道路工程", "路基施工", 150],
            [None,       None,       "路面施工", 120],
            [None,       "围墙工程", "基础施工", 80],
            [None,       None,       "墙体砌筑", 60],
        ]
        for i, row_data in enumerate(data, start=2):
            for j, val in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        rules = [
            {"user_header": "单位工程", "system_field": "unit"},
            {"user_header": "分部工程", "system_field": "sub_unit"},
            {"user_header": "分项工程", "system_field": "item"},
        ]
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=1, rules=rules)
        _fill_merged_cells(raw_rows)

        # 单位工程：前4行 → 粮仓主体
        for i in range(4):
            assert raw_rows[i]["单位工程"] == "粮仓主体"
        # 单位工程：后4行 → 附属工程
        for i in range(4, 8):
            assert raw_rows[i]["单位工程"] == "附属工程"

        # 分部工程：第1-2行 → 基础工程；第3-4行 → 主体结构
        assert raw_rows[0]["分部工程"] == "基础工程"
        assert raw_rows[1]["分部工程"] == "基础工程"
        assert raw_rows[2]["分部工程"] == "主体结构"
        assert raw_rows[3]["分部工程"] == "主体结构"

        # 分项工程不补全（None 不被列为合并范围——但当前实现会补全）
        # 注意：当前 _fill_merged_cells 对所有 None 列都补全，
        # 包括分项工程列。这是正确行为还是 bug 取决于业务场景。
        # 对于真正的分项工程（每行都不同），不应该被合并补全。
        # 但目前实现确实会补全，这是已知的局限。

    def test_empty_rows_filled(self):
        """空行也能通过合并补全获得上一行的值"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        headers = ["类别", "名称", "金额"]
        for j, h in enumerate(headers, start=1):
            ws.cell(row=1, column=j, value=h)

        data = [
            ["建安工程", "浅圆仓", 500],
            [None,       None,    300],   # 全空行 → 全部继承上一行
        ]
        for i, row_data in enumerate(data, start=2):
            for j, val in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        rules = [
            {"user_header": "类别", "system_field": "category"},
            {"user_header": "名称", "system_field": "name"},
            {"user_header": "金额", "system_field": "amount"},
        ]
        raw_rows = _read_xlsx_rows(xlsx_bytes, sheet_index=0, header_row=1, rules=rules)
        _fill_merged_cells(raw_rows)

        # 第2行全部继承第1行
        assert raw_rows[1]["类别"] == "建安工程"
        assert raw_rows[1]["名称"] == "浅圆仓"
        assert raw_rows[1]["金额"] == 300  # 注意：金额也继承了！这是当前实现的局限
